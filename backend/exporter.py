import csv
import io
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, fills

logger = logging.getLogger(__name__)

IMDB_COLUMNS = [
    "record_id",
    "barcode",
    "category_type",
    "segment_type",
    "manufacturer",
    "brand",
    "product_name",
    "weight_unit",
    "packaging_type",
    "country_origin",
    "marketing_message",
    "confidence",
    "flagged_for_review",
    "potential_duplicate",
    "notes",
    "source_filename",
]

_SKIP_IN_LOOP = {"confidence", "flagged_for_review", "potential_duplicate", "notes", "source_filename", "record_id"}


def records_to_dataframe(records: List[Dict]) -> pd.DataFrame:
    rows = []
    for rec in records:
        row = {col: rec.get(col, "") for col in IMDB_COLUMNS if col not in _SKIP_IN_LOOP}
        row["record_id"] = rec.get("record_id", "")
        row["confidence"] = rec.get("confidence", 100)
        row["flagged_for_review"] = "Yes" if rec.get("flagged_for_review") else "No"
        row["potential_duplicate"] = "Yes" if rec.get("potential_duplicate") else "No"
        row["notes"] = rec.get("notes", "")
        row["source_filename"] = rec.get("source_filename", "")
        rows.append(row)
    return pd.DataFrame(rows, columns=IMDB_COLUMNS)


def export_to_excel(records: List[Dict], filename: str = None) -> bytes:
    df = records_to_dataframe(records)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="IMDB Import")
        worksheet = writer.sheets["IMDB Import"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        # Row colouring: orange = flagged for review, yellow = potential duplicate
        flag_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        dup_fill  = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
        flag_col = IMDB_COLUMNS.index("flagged_for_review") + 1
        dup_col  = IMDB_COLUMNS.index("potential_duplicate") + 1
        for row in worksheet.iter_rows(min_row=2):
            is_flagged = row[flag_col - 1].value == "Yes"
            is_dup     = row[dup_col - 1].value == "Yes"
            if is_dup:
                for cell in row:
                    cell.fill = dup_fill
            elif is_flagged:
                for cell in row:
                    cell.fill = flag_fill

        for column in worksheet.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            worksheet.column_dimensions[col_letter].width = min(max_length + 4, 50)
    return output.getvalue()


def export_to_csv(records: List[Dict]) -> str:
    df = records_to_dataframe(records)
    return df.to_csv(index=False)
